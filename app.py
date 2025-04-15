from flask import Flask, render_template_string, request, redirect, url_for, send_from_directory
import os
import pdfplumber
import openpyxl
from werkzeug.utils import secure_filename

app = Flask(__name__)

# Define o diretório para uploads e o caminho para o arquivo Excel
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')
EXCEL_FILE_NAME = 'base_dados1.xlsx Final.xlsx'
EXCEL_FILE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILE_NAME)
ALLOWED_EXTENSIONS = {'pdf'}

# Cria o diretório de uploads, caso não exista
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Configurações do Flask
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # Limite de 16MB para o upload de arquivos

# Template HTML para o formulário de upload com a identidade visual do Sicoob
html_template = """
<!doctype html>
<html lang="pt-br">
<head>
  <meta charset="utf-8">
  <title>Automação da Prestação de Contas </title>
  <link href="https://fonts.googleapis.com/css2?family=Roboto:wght@400;500;700&display=swap" rel="stylesheet">
  <style>
    body {
      font-family: 'Roboto', sans-serif;
      background-image: url("{{ url_for('static', filename='background.jpg') }}");
      background-size: cover;
      background-repeat: no-repeat;
      background-position: center center;
      background-attachment: fixed;
      color: #333;
      margin: 0;
      padding: 0;
    }
    header {
      background-color: rgba(0, 174, 157, 0.9);
      color: white;
      padding: 30px 20px;
      text-align: center;
      box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
    }
    header img {
      width: 150px;
      margin-bottom: 15px;
    }
    h1 {
      font-size: 2rem;
      margin: 0;
      letter-spacing: 1px;
    }
    .container {
      width: 90%;
      max-width: 800px;
      margin: 30px auto;
      padding: 20px;
      background-color: rgba(255, 255, 255, 0.95);
      box-shadow: 0 6px 12px rgba(0, 0, 0, 0.1);
      border-radius: 8px;
    }
    h2 {
      font-size: 1.5rem;
      margin-bottom: 15px;
      text-align: center;
    }
    .message {
      margin: 10px 0;
      padding: 10px;
      border-radius: 5px;
      text-align: center;
      font-weight: 500;
    }
    .message.success {
      background-color: #d4edda;
      color: #155724;
    }
    .message.error {
      background-color: #f8d7da;
      color: #721c24;
    }
    form {
      display: flex;
      flex-direction: column;
      align-items: center;
    }
    input[type="file"] {
      padding: 10px;
      margin-bottom: 20px;
      border: 1px solid #ddd;
      border-radius: 5px;
      width: 100%;
      max-width: 350px;
    }
    input[type="submit"] {
      padding: 12px 25px;
      background-color: #2a8c45;
      color: white;
      border: none;
      border-radius: 5px;
      cursor: pointer;
      font-size: 1rem;
      transition: background-color 0.3s ease;
    }
    input[type="submit"]:hover {
      background-color: #00AE9D;
    }
    a {
      color: #00AE9D;
      text-decoration: none;
      font-weight: 500;
      margin-top: 15px;
      display: block;
      text-align: center;
    }
    a:hover {
      text-decoration: underline;
    }
    footer {
      background-color: #f1f1f1;
      padding: 20px;
      text-align: center;
      font-size: 0.9rem;
      color: #777;
    }
    @media (max-width: 600px) {
      .container {
        width: 100%;
        margin: 15px;
      }
      h1 {
        font-size: 1.5rem;
      }
    }
  </style>
</head>
<body>
  <header>
    <img src="{{ url_for('static', filename='logo_sicoob.png') }}" alt="Logo Sicoob">
    <h1>Automação de Prestação de contas</h1>
  </header>

  <div class="container">
    <h2>Selecione o arquivo PDF para ser processado</h2>

    {% with messages = get_flashed_messages(with_categories=true) %}
      {% if messages %}
        {% for category, message in messages %}
          <div class="message {{ category }}">{{ message }}</div>
        {% endfor %}
      {% endif %}
    {% endwith %}
    
    <form method="post" enctype="multipart/form-data" action="{{ url_for('upload_file') }}">
      <input type="file" name="pdf_file" accept=".pdf" required>
      <input type="submit" value="Processar PDF">
    </form>

    {% if excel_exists %}
      <p><a href="{{ url_for('download_excel') }}">Download do Excel Atualizado</a></p>
    {% endif %}
  </div>

  <footer>
    <p>&copy; 2025 Sicoob Nova Central - Todos os direitos reservados</p>
  </footer>
</body>
</html>
"""

# Função para verificar se o arquivo tem a extensão permitida
def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Rota para a página inicial com o formulário de upload
@app.route('/', methods=['GET'])
def index():
    excel_exists = os.path.exists(EXCEL_FILE_PATH)
    return render_template_string(html_template, excel_exists=excel_exists)

# Rota para processar o upload do arquivo PDF
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'pdf_file' not in request.files:
        return redirect(url_for('index'))

    file = request.files['pdf_file']

    if file.filename == '':
        return redirect(url_for('index'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(pdf_path)

        try:
            if os.path.exists(EXCEL_FILE_PATH):
                workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
            else:
                workbook = openpyxl.Workbook()
                if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
                    del workbook["Sheet"]

            sheet_name = "Solicitação TXT"
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                max_r = sheet.max_row
                max_c = sheet.max_column
                merged_ranges_coords = {cell for mr in sheet.merged_cells.ranges for cell in mr.coord.split(':') if ':' in mr.coord for cell in openpyxl.utils.rows_from_range(mr.coord)}
                if not merged_ranges_coords: # Handle single cell merges or ranges without ':'
                    merged_ranges_coords = {cell.coord for mr in sheet.merged_cells.ranges for cell in mr}


                for r in range(1, max_r + 1):
                     for c in range(1, max_c + 1):
                         cell = sheet.cell(row=r, column=c)
                         if not isinstance(cell, openpyxl.cell.MergedCell) and cell.coordinate not in merged_ranges_coords:
                             cell.value = None
            else:
                sheet = workbook.create_sheet(sheet_name)

            # Processa o PDF e escreve os dados na planilha
            with pdfplumber.open(pdf_path) as pdf:
                merged_ranges_coords = {cell for mr in sheet.merged_cells.ranges for cell in mr.coord.split(':') if ':' in mr.coord for cell in openpyxl.utils.rows_from_range(mr.coord)}
                if not merged_ranges_coords: # Handle single cell merges or ranges without ':'
                    merged_ranges_coords = {cell.coord for mr in sheet.merged_cells.ranges for cell in mr}

                for page in pdf.pages:
                    table = page.extract_table()
                    if table:
                        for row_idx, row_data in enumerate(table, start=1):
                             for col_idx, cell_data in enumerate(row_data, start=1):
                                 target_cell = sheet.cell(row=row_idx, column=col_idx)
                                 if not isinstance(target_cell, openpyxl.cell.MergedCell) and target_cell.coordinate not in merged_ranges_coords:
                                     target_cell.value = cell_data

            # Salva o arquivo Excel atualizado
            workbook.save(EXCEL_FILE_PATH)

        except Exception as e:
            print(f"Erro: {e}")
        finally:
            if os.path.exists(pdf_path):
                os.remove(pdf_path)

        return redirect(url_for('index'))

    else:
        return redirect(url_for('index'))

# Rota para download do arquivo Excel atualizado
@app.route('/download')
def download_excel():
    try:
        return send_from_directory(
            directory=os.path.dirname(EXCEL_FILE_PATH),
            path=EXCEL_FILE_NAME,
            as_attachment=True
        )
    except FileNotFoundError:
        return redirect(url_for('index'))

# Roda o servidor Flask em todas as interfaces de rede (0.0.0.0)
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
